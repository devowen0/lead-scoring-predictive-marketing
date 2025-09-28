import pandas as pd
from datetime import datetime
import os
import unicodedata
import pyperclip
from openpyxl import load_workbook

# --- Ask user for date input ---
while True:
    user_input = input("Enter the date to search for (format: YYYY-MM-DD, preferably between 2025-09-27 to 2025-11-01): ").strip()
    try:
        selected_date = datetime.strptime(user_input, "%Y-%m-%d").date()
        break
    except ValueError:
        print("❌ Invalid format. Please enter the date as YYYY-MM-DD (example: 2025-01-01).")

# Load the Excel file with pandas for data access
file_path = "demo_leads_scored.xlsx"
df = pd.read_excel(file_path)

# Columns to check
date_columns = [
    "Education Date", "Feedback Date", "Welcome Date",
    "Promo 1 Date", "Promo 2 Date", "Promo 3 Date",
    "Promo 4 Date", "Promo 5 Date", "Promo 6 Date", "Promo 7 Date"
]

# Base folder
base_folder = "messages"

# Map date columns to subfolders
column_folder_map = {
    "Education Date": "educational",
    "Feedback Date": "feedback",
    "Welcome Date": "welcome",
    "Promo 1 Date": "promotion",
    "Promo 2 Date": "promotion",
    "Promo 3 Date": "promotion",
    "Promo 4 Date": "promotion",
    "Promo 5 Date": "promotion",
    "Promo 6 Date": "promotion",
    "Promo 7 Date": "promotion"
}

# Map industry to subfolders
industry_folder_map = {
    "Bygg": "Bygg",
    "Detaljhandel": "Detaljhandel",
    "IT-tjänster": "it",
    "Konsult": "konsult",
    "Marknadsföring": "Marknadsföring"
}

# Normalize placeholders to NFC (so [Förnamn] and [Förnamn] match)
PLACEHOLDERS = [unicodedata.normalize("NFC", "[First Name]"),
                unicodedata.normalize("NFC", "[Förnamn]")]

def replace_placeholders(text, first_name):
    """Normalize text to NFC and replace placeholders with the first name."""
    if not isinstance(text, str):
        return text
    normalized = unicodedata.normalize("NFC", text)
    for ph in PLACEHOLDERS:
        normalized = normalized.replace(ph, first_name)
    return normalized

# Pre-scan to count total matches
total_matches = 0
for _, row in df.iterrows():
    for col in date_columns:
        cell = row.get(col)
        if pd.isna(cell) or str(cell).strip().upper() == "N/A":
            continue
        try:
            cell_date = pd.to_datetime(cell).date() if not isinstance(cell, pd.Timestamp) else cell.date()
        except:
            continue
        if cell_date == selected_date:
            total_matches += 1
            break

print(f"\n🔎 Total people with a date matching {selected_date}: {total_matches}\n")

# Load the workbook for single-cell updates
wb = load_workbook(file_path)
ws = wb.active  # Assuming the first sheet is the correct one

# Iterate through each row
for row_index, row in df.iterrows():
    matched_columns = []

    for col in date_columns:
        cell = row.get(col)
        if pd.isna(cell) or str(cell).strip().upper() == "N/A":
            continue
        try:
            cell_date = pd.to_datetime(cell).date() if not isinstance(cell, pd.Timestamp) else cell.date()
        except:
            continue
        if cell_date == selected_date:
            matched_columns.append(col)

    if matched_columns:
        language = row.get("Swedish/English", "No Preference Provided")
        email = row.get("Email", "No Email Provided")
        first_name = row.get("First Name", "")
        industry_raw = row.get("Industry", "")

        # Normalize industry name to NFC
        industry = unicodedata.normalize("NFC", str(industry_raw))

        print(f"📧 Email: {email}")

        # Determine language folder
        if language.lower().startswith("english"):
            language_folder = "English"
        elif language.lower().startswith("swedish") or language.lower().startswith("svenska"):
            language_folder = "Svenska"
        else:
            language_folder = ""

        # Determine industry folder
        industry_folder = industry_folder_map.get(industry, "")

        for match_col in matched_columns:
            folder_type = column_folder_map.get(match_col, "")
            folder_path = os.path.join(base_folder, language_folder, folder_type, industry_folder)

            if not os.path.exists(folder_path):
                print(f"⚠️ Folder not found: {folder_path}")
                continue

            txt_files = [f for f in os.listdir(folder_path) if f.endswith(".txt")]
            selected_file = None

            if "Promo" in match_col:
                promo_number = match_col.split()[1]
                for f in txt_files:
                    if f.startswith(promo_number):
                        selected_file = f
                        break
            else:
                if txt_files:
                    selected_file = txt_files[0]

            if not selected_file:
                print(f"⚠️ No .txt file found in {folder_path} for {match_col}")
                continue

            # Read and process content
            file_path_full = os.path.join(folder_path, selected_file)
            with open(file_path_full, "r", encoding="utf-8") as file:
                content = file.read()

            content = replace_placeholders(content, first_name)

            # Determine and process subject
            subject = os.path.splitext(selected_file)[0]
            subject = replace_placeholders(subject, first_name)

            if "promotion" in folder_path.lower() and subject:
                subject = subject[1:]

            # Print subject and content
            print(f"\n📌 Subject: {subject}\n")
            print(content)
            print("\n" + "-" * 50 + "\n")

            # Clipboard workflow: email → subject → body
            pyperclip.copy(email)
            input("📋 Email copied to clipboard. Press Enter to copy subject...")
            pyperclip.copy(subject)
            input("📋 Subject copied to clipboard. Press Enter to copy email body...")
            pyperclip.copy(content)
            input("📋 Email body copied to clipboard. Press Enter when ready to confirm sending...")

            # Prompt until user types "yes"
            while True:
                user_input = input("Type 'yes' to confirm you've sent the email: ").strip().lower()
                if user_input == "yes":
                    # Find the column number in Excel
                    col_number = None
                    for idx, header in enumerate(ws[1], start=1):
                        if header.value == match_col:
                            col_number = idx
                            break
                    if col_number is not None:
                        excel_row = row_index + 2  # Adjust for header row
                        ws.cell(row=excel_row, column=col_number, value="DONE")
                        wb.save(file_path)
                        print(f"✅ Updated cell {match_col} to 'DONE' for {email}\n")
                    else:
                        print(f"⚠️ Could not find column {match_col} in Excel to update.")
                    break
